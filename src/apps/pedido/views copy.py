# apps/pedido/views.py

from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum
from django.db.models.functions import TruncMonth

from collections import OrderedDict
from datetime import date as datetime_date
from io import BytesIO

# --- CORREÇÃO NA SEÇÃO DE IMPORTS ---
# Precisamos importar a classe 'Workbook' diretamente do openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
# --------------------------------------

from .models import Pedido
from .templatetags.pedido_extras import soma_qt_mt

# ==============================================================================
# 1. CONSTANTES E FUNÇÕES AUXILIARES CENTRALIZADAS
# ==============================================================================

MESES_DO_ANO = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}

PALAVRAS_CHAVE_ARTIGOS = [
    "LATEGO", "NAPA", "NOBUCK", "VAQUETA ASA", "VAQUETA RELAX - SEGURANCA", 
    "VAQUETA", "CAMURÇA", "CHAMOIS", "FORRO", "BOX", "COURO PELE", "PELE", 
    "PULL UP", "CRAZY HORSE", "FLOATER", "RASPA", "ARTIGO VAZIO"
]

def get_grupo_artigo(nome_artigo):
    """
    Função única e centralizada para categorizar um artigo com base em palavras-chave.
    """
    if not nome_artigo or not nome_artigo.strip():
        return "ARTIGO VAZIO"
    
    nome_artigo_upper = nome_artigo.upper()
    
    for palavra_chave in PALAVRAS_CHAVE_ARTIGOS:
        if palavra_chave in nome_artigo_upper:
            return palavra_chave
    
    return "OUTROS TIPOS"

def get_pedidos_base_queryset(ids: list):
    """
    Função base que retorna o queryset inicial de pedidos com otimizações de consulta.
    """
    if not ids:
        return Pedido.objects.none()
    
    return Pedido.objects.filter(id__in=ids).prefetch_related(
        'requisicao_links__requisicao__fluxos__processo'
    )

def anexa_processos_consolidados(pedidos: list):
    """
    Anexa a informação de processos consolidados a cada objeto de pedido em uma lista.
    """
    for pedido in pedidos:
        processos_atuais_req_nomes = set()
        for link in pedido.requisicao_links.all():
            try:
                processo_atual_req = link.requisicao.get_processo_atual()
                if processo_atual_req and processo_atual_req != "Não definido":
                    processos_atuais_req_nomes.add(processo_atual_req)
            except Exception:
                continue
        pedido.processos_consolidados_requisitantes = sorted(list(processos_atuais_req_nomes), key=lambda p: p.dt_processo, reverse=True)


# ==============================================================================
# 2. VIEWS REATORADAS
# ==============================================================================

@staff_member_required
def imprimir_pedido_view(request):
    """
    Gera o relatório DETALHADO, com layout de cards, agrupado por Artigo e Cliente.
    """
    ids_str = request.GET.get("ids", "")
    ids = [int(i) for i in ids_str.split(",") if i.strip().isdigit()] if ids_str else []

    pedidos_selecionados = get_pedidos_base_queryset(ids).order_by('artigo', 'cliente')
    anexa_processos_consolidados(pedidos_selecionados)

    totais_qt_por_grupo = {}
    # Este loop é necessário para usar a função get_grupo_artigo em Python
    for pedido in pedidos_selecionados:
        grupo = get_grupo_artigo(pedido.artigo)
        totais_qt_por_grupo[grupo] = totais_qt_por_grupo.get(grupo, 0) + (pedido.quantidade or 0)
    
    totais_mensais_query = pedidos_selecionados.filter(dt_programada__isnull=False).annotate(
        mes_ano=TruncMonth('dt_programada')
    ).values('artigo', 'mes_ano').annotate(total_qt_mes=Sum('quantidade')).order_by('artigo', 'mes_ano')

    totais_artigo_por_mes = {}
    for item in totais_mensais_query:
        grupo_artigo = get_grupo_artigo(item['artigo'])
        mes_display = f"{MESES_DO_ANO[item['mes_ano'].month]}/{item['mes_ano'].year}"
        totais_artigo_por_mes.setdefault(grupo_artigo, OrderedDict())
        totais_artigo_por_mes[grupo_artigo][mes_display] = totais_artigo_por_mes[grupo_artigo].get(mes_display, 0) + item['total_qt_mes']

    pedidos_agrupados_final = OrderedDict()
    totais_prod_por_grupo = {}

    for pedido in pedidos_selecionados:
        grupo_artigo = get_grupo_artigo(pedido.artigo)
        cliente = pedido.cliente or "CLIENTE NÃO ESPECIFICADO"
        
        pedidos_agrupados_final.setdefault(grupo_artigo, OrderedDict())
        pedidos_agrupados_final[grupo_artigo].setdefault(cliente, [])
        pedidos_agrupados_final[grupo_artigo][cliente].append(pedido)

        soma_producao_pedido = soma_qt_mt(pedido.requisicao_links.all())
        totais_prod_por_grupo[grupo_artigo] = totais_prod_por_grupo.get(grupo_artigo, 0) + soma_producao_pedido
        
    totais_por_artigo_final = {}
    for grupo, total_qt in totais_qt_por_grupo.items():
        totais_por_artigo_final[grupo] = {
            'total_qt': total_qt,
            'total_prod': totais_prod_por_grupo.get(grupo, 0)
        }

    context = {
        "pedidos_agrupados_por_artigo_e_cliente": pedidos_agrupados_final,
        "totais_por_artigo": totais_por_artigo_final,
        "totais_artigo_por_mes": totais_artigo_por_mes,
        "data_atual": timezone.now()
    }
    return render(request, "pedido/impressao.html", context)


@staff_member_required
def imprimir_pedido_data_view(request):
    """
    Gera o relatório agrupado por MÊS e CLIENTE, com subtotais e contagem de dias até o prazo.
    """
    from datetime import date

    ids_str = request.GET.get("ids", "")
    output_format = request.GET.get("formato", "html")
    ids = [int(i) for i in ids_str.split(",") if i.strip().isdigit()] if ids_str else []

    pedidos_selecionados = Pedido.objects.none()
    if ids:
        pedidos_selecionados = Pedido.objects.filter(id__in=ids).order_by(
            'dt_programada', 'cliente', 'artigo', 'unidade_medida'
        )

    # Adiciona campo calculado de dias restantes/atraso
    from datetime import datetime, date
    hoje = date.today()
    for p in pedidos_selecionados:
        if p.dt_programada:
            # Se for datetime, converte para date antes de subtrair
            data_prazo = p.dt_programada.date() if isinstance(p.dt_programada, datetime) else p.dt_programada
            p.dias_restantes = (data_prazo - hoje).days
        else:
            p.dias_restantes = None

    # Totais por mês (mantido)
    from django.db.models.functions import TruncMonth
    from django.db.models import Sum
    totais_mensais_query = pedidos_selecionados.filter(dt_programada__isnull=False).annotate(
        mes=TruncMonth('dt_programada')
    ).values('mes', 'unidade_medida').annotate(total_quantidade=Sum('quantidade')).order_by('mes', 'unidade_medida')

    totais_mensais_dict = {}
    for total in totais_mensais_query:
        mes_chave = f"{total['mes'].year}-{total['mes'].month:02d}"
        unidade = total['unidade_medida'] or 'N/D'
        totais_mensais_dict.setdefault(mes_chave, {})[unidade] = total['total_quantidade']

    # Agrupamento por mês → cliente → artigo
    from collections import OrderedDict
    pedidos_agrupados = OrderedDict()
    for pedido in pedidos_selecionados:
        if not pedido.dt_programada:
            continue

        mes_chave = f"{pedido.dt_programada.year}-{pedido.dt_programada.month:02d}"
        cliente = pedido.cliente or "CLIENTE NÃO INFORMADO"
        artigo = pedido.artigo or "ARTIGO NÃO INFORMADO"

        if mes_chave not in pedidos_agrupados:
            pedidos_agrupados[mes_chave] = {
                'clientes': OrderedDict(),
                'subtotais': totais_mensais_dict.get(mes_chave, {}),
                'mes_display': f"{MESES_DO_ANO[pedido.dt_programada.month]} - {pedido.dt_programada.year}"
            }

        if cliente not in pedidos_agrupados[mes_chave]['clientes']:
            pedidos_agrupados[mes_chave]['clientes'][cliente] = {
                'artigos': OrderedDict(),
                'subtotais_cliente': {}
            }

        if artigo not in pedidos_agrupados[mes_chave]['clientes'][cliente]['artigos']:
            pedidos_agrupados[mes_chave]['clientes'][cliente]['artigos'][artigo] = []

        pedidos_agrupados[mes_chave]['clientes'][cliente]['artigos'][artigo].append(pedido)

        unidade = pedido.unidade_medida or "N/D"
        subtotal_cliente = pedidos_agrupados[mes_chave]['clientes'][cliente]['subtotais_cliente']
        subtotal_cliente[unidade] = subtotal_cliente.get(unidade, 0) + (pedido.quantidade or 0)

    # Exportação Excel (mantida)
    if output_format == 'excel':
        totais_gerais_dict = {
            item['unidade_medida'] or 'N/D': item['total_geral']
            for item in pedidos_selecionados.values('unidade_medida').annotate(total_geral=Sum('quantidade'))
        }
        excel_buffer = gerar_excel_data_view(pedidos_agrupados, totais_gerais_dict)
        from django.http import HttpResponse
        response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_pedidos_por_data.xlsx"'
        return response

    from django.utils import timezone
    context = {
        "pedidos_agrupados": pedidos_agrupados,
        "data_atual": timezone.now()
    }
    return render(request, "pedido/impressao_data.html", context)



# ==============================================================================
# 3. FUNÇÕES DE EXPORTAÇÃO PARA EXCEL
# ==============================================================================

def gerar_excel_data_view(pedidos_agrupados, totais_gerais):
    """
    Função dedicada a gerar a planilha para o relatório agrupado por DATA.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pedidos por Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    group_font = Font(bold=True, size=12)
    total_font = Font(bold=True)
    
    headers = ["Contrato", "Cód. Interno", "Cliente", "Artigo", "Quantidade", "Unidade", "Prazo"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for mes, dados_mes in pedidos_agrupados.items():
        sheet.append([])
        group_cell = sheet.cell(row=sheet.max_row, column=1, value=dados_mes['mes_display'])
        group_cell.font = group_font
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=len(headers))
        
        for pedido in dados_mes['items']:
            row_data = [
                pedido.nr_contract, pedido.nr_pedido_interno, pedido.cliente,
                pedido.artigo, pedido.quantidade, pedido.unidade_medida,
                pedido.dt_programada.strftime('%d/%m/%Y') if pedido.dt_programada else 'N/A'
            ]
            sheet.append(row_data)
        
        for i, (unidade, total) in enumerate(dados_mes['subtotais'].items()):
            label = "Subtotais:" if i == 0 else ""
            sheet.append(["", "", "", label, total, f"({unidade})", ""])
            for cell in sheet[sheet.max_row]:
                cell.font = total_font
    
    sheet.append([])
    for i, (unidade, total) in enumerate(totais_gerais.items()):
        label = "TOTAIS GERAIS:" if i == 0 else ""
        sheet.append(["", "", "", label, total, f"({unidade})", ""])
        for cell in sheet[sheet.max_row]:
            cell.font = total_font
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
    for i, column_cells in enumerate(sheet.columns):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(i + 1)].width = max_length + 2

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer