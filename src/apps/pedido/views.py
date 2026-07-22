# apps/pedido/views.py

from __future__ import annotations
from dotenv import load_dotenv, find_dotenv
load_dotenv(find_dotenv(), override=False)

import json
from datetime import timedelta
from io import BytesIO
from collections import OrderedDict


from django.views.decorators.http import require_POST
from django.utils.dateparse import parse_date

from django.views.decorators.http import require_GET
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Pedido,
    Veiculo,
    Embarque,
    ItemEmbarque,
    TipoVeiculo,
    Transportadora,
    PrevisaoEmbarque,      # novo
    PrevisaoItemEmbarque,  # novo
)

import os
import pymssql
from .selectpedidos import SelectPedidos

PES2_TO_M2_DIVISOR = 10.764


# ==============================================================================
# 1) CONSTANTES / FUNÇÕES AUXILIARES (relatórios)
# ==============================================================================

MESES_DO_ANO = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}

PALAVRAS_CHAVE_ARTIGOS = [
    "LATEGO", "NAPA", "NOBUCK", "VAQUETA ASA", "VAQUETA RELAX - SEGURANCA",
    "VAQUETA", "CAMURÇA", "CHAMOIS", "FORRO", "BOX", "COURO PELE", "PELE",
    "PULL UP", "CRAZY HORSE", "FLOATER", "RASPA", "ARTIGO VAZIO"
]


def get_grupo_artigo(nome_artigo: str | None) -> str:
    if not nome_artigo or not nome_artigo.strip():
        return "ARTIGO VAZIO"
    nome_artigo_upper = nome_artigo.upper()
    for palavra_chave in PALAVRAS_CHAVE_ARTIGOS:
        if palavra_chave in nome_artigo_upper:
            return palavra_chave
    return "OUTROS TIPOS"


def get_pedidos_base_queryset(ids: list[int]):
    if not ids:
        return Pedido.objects.none()
    return (
        Pedido.objects.filter(id__in=ids)
        .prefetch_related("requisicao_links__requisicao__fluxos__processo", "historico_embarques__embarque__veiculo")
    )


def anexa_processos_consolidados(pedidos: list[Pedido]):
    for pedido in pedidos:
        processos_atuais_req_nomes = set()
        for link in pedido.requisicao_links.all():
            try:
                processo_atual_req = link.requisicao.get_processo_atual()
                if processo_atual_req and processo_atual_req != "Não definido":
                    processos_atuais_req_nomes.add(processo_atual_req)
            except Exception:
                continue
        pedido.processos_consolidados_requisitantes = sorted(
            list(processos_atuais_req_nomes), key=lambda p: p.dt_processo, reverse=True
        )



# =============================================================================
# DASHBOARD
# =============================================================================


@staff_member_required
def dashboard_view(request, *args, **kwargs):   
    agora = timezone.now()
    trinta_dias_atras = agora - timedelta(days=30)

    # ========================================================
    # 1. MÉTRICAS DE PEDIDOS (COMERCIAL)
    # ========================================================
    pedidos_abertos = Pedido.objects.filter(fechado=False)
    
    # Total em carteira (m²)
    total_carteira_m2 = pedidos_abertos.aggregate(total=Sum('quantidade'))['total'] or 0
    
    # Total aguardando embarque (saldo > 0)
    # Como saldo_a_entregar é @property, calculamos no Python
    pedidos_pendentes = [p for p in pedidos_abertos if p.saldo_a_entregar > 0]
    total_pendente_m2 = sum(p.saldo_a_entregar for p in pedidos_pendentes)
    
    # Atrasados (Programação já passou)
    pedidos_atrasados = [
        p for p in pedidos_pendentes 
        if p.dt_programada and p.dt_programada.date() < agora.date()
    ]

    # ========================================================
    # 2. MÉTRICAS DE REQUISIÇÕES (PRODUÇÃO)
    # ========================================================
    from src.apps.fluxo.models import Requisicao # Ajuste o import se estiver em outra pasta
    
    # Requisições em aberto (não encerradas) nos últimos 30 dias
    reqs_recentes = Requisicao.objects.filter(dt_requisicao__gte=trinta_dias_atras)
    
    total_reqs_abertas = reqs_recentes.filter(encerrado=False).count()
    
    # Produção em andamento (m² e peças)
    producao_m2 = reqs_recentes.filter(encerrado=False).aggregate(total=Sum('m2'))['total'] or 0
    producao_pecas = reqs_recentes.filter(encerrado=False).aggregate(total=Sum('quantidade'))['total'] or 0

    # ========================================================
    # 3. MÉTRICAS DE LOGÍSTICA (EMBARQUES)
    # ========================================================
    # Embarques em andamento e finalizados no mês
    embarques_mes = Embarque.objects.filter(data_embarque__month=agora.month, data_embarque__year=agora.year)
    
    embarques_abertos = embarques_mes.filter(finalizado=False).count()
    embarques_finalizados = embarques_mes.filter(finalizado=True).count()
    
    # Volume embarcado (m²) no mês atual
    volume_embarcado_mes = ItemEmbarque.objects.filter(
        embarque__data_embarque__month=agora.month,
        embarque__data_embarque__year=agora.year
    ).aggregate(total=Sum('metragem_embarcada'))['total'] or 0

    context = {
        # Pedidos
        'total_carteira_m2': total_carteira_m2,
        'total_pendente_m2': total_pendente_m2,
        'qtd_pedidos_atrasados': len(pedidos_atrasados),
        'm2_atrasado': sum(p.saldo_a_entregar for p in pedidos_atrasados),
        
        # Produção
        'total_reqs_abertas': total_reqs_abertas,
        'producao_m2': producao_m2,
        'producao_pecas': producao_pecas,
        
        # Logística
        'embarques_abertos': embarques_abertos,
        'embarques_finalizados': embarques_finalizados,
        'volume_embarcado_mes': volume_embarcado_mes,
        
        # Auxiliares
        'mes_atual': agora.strftime("%B/%Y"),
    }

    return render(request, "pedido/dashboard.html", context)





# ==============================================================================
# 2) RELATÓRIOS (mantidos)
# ==============================================================================

@staff_member_required
def imprimirpedidoview(request):
    ids_str = request.GET.get("ids", "")
    ids = [int(i) for i in ids_str.split(",") if i.strip().isdigit()] if ids_str else []

    pedidos_selecionados = get_pedidos_base_queryset(ids).order_by("artigo", "cliente")
    anexa_processos_consolidados(pedidos_selecionados)

    totais_qt_por_grupo = {}
    totais_saldo_por_grupo = {}

    for pedido in pedidos_selecionados:
        grupo = get_grupo_artigo(pedido.artigo)
        totais_qt_por_grupo[grupo] = totais_qt_por_grupo.get(grupo, 0) + float(pedido.quantidade or 0)
        totais_saldo_por_grupo[grupo] = totais_saldo_por_grupo.get(grupo, 0) + float(pedido.saldo_a_entregar or 0)

    totais_mensais_query = (
        pedidos_selecionados.filter(dt_programada__isnull=False)
        .annotate(mes_ano=TruncMonth("dt_programada"))
        .values("artigo", "mes_ano")
        .annotate(total_qt_mes=Sum("quantidade"))
        .order_by("artigo", "mes_ano")
    )

    totais_artigo_por_mes = {}
    for item in totais_mensais_query:
        grupo_artigo = get_grupo_artigo(item["artigo"])
        mes_display = f"{MESES_DO_ANO[item['mes_ano'].month]}/{item['mes_ano'].year}"
        totais_artigo_por_mes.setdefault(grupo_artigo, OrderedDict())
        totais_artigo_por_mes[grupo_artigo][mes_display] = totais_artigo_por_mes[grupo_artigo].get(mes_display, 0) + (
            item["total_qt_mes"] or 0
        )

    pedidos_agrupados_final = OrderedDict()
    totais_prod_por_grupo = {}

    def soma_producao(_links):
        return 0

    for pedido in pedidos_selecionados:
        grupo_artigo = get_grupo_artigo(pedido.artigo)
        cliente = pedido.cliente or "CLIENTE NÃO ESPECIFICADO"
        pedidos_agrupados_final.setdefault(grupo_artigo, OrderedDict())
        pedidos_agrupados_final[grupo_artigo].setdefault(cliente, [])
        pedidos_agrupados_final[grupo_artigo][cliente].append(pedido)

        totais_prod_por_grupo[grupo_artigo] = totais_prod_por_grupo.get(grupo_artigo, 0) + soma_producao(
            pedido.requisicao_links.all()
        )

    totais_por_artigo_final = {}
    for grupo, total_qt in totais_qt_por_grupo.items():
        totais_por_artigo_final[grupo] = {
            "total_qt": total_qt,
            "total_prod": totais_prod_por_grupo.get(grupo, 0),
            "total_saldo": totais_saldo_por_grupo.get(grupo, 0),
        }

    context = {
        "pedidos_agrupados_por_artigo_e_cliente": pedidos_agrupados_final,
        "totais_por_artigo": totais_por_artigo_final,
        "totais_artigo_por_mes": totais_artigo_por_mes,
        "data_atual": timezone.now(),
    }
    return render(request, "pedido/impressao.html", context)


@staff_member_required
def imprimirpedidodataview(request):
    output_format = request.GET.get("formato", "html")

    pedidos_selecionados = (
        Pedido.objects.filter(dt_programada__isnull=False)
        .filter(Q(fechado=False) | Q(fechado__isnull=True))
        .prefetch_related("historico_embarques")
        .order_by("dt_programada", "cliente", "unidade_medida")
    )

    ids_str = request.GET.get("ids", "")
    if ids_str:
        ids = [int(i) for i in ids_str.split(",") if i.strip().isdigit()]
        if ids:
            pedidos_selecionados = pedidos_selecionados.filter(id__in=ids)

    totais_mensais_query = (
        pedidos_selecionados.annotate(mes=TruncMonth("dt_programada"))
        .values("mes", "unidade_medida")
        .annotate(total_quantidade=Sum("quantidade"))
        .order_by("mes", "unidade_medida")
    )

    totais_mensais_dict = {}
    for total in totais_mensais_query:
        mes_chave = f"{total['mes'].year}-{total['mes'].month:02d}"
        unidade = total["unidade_medida"] or "N/D"
        totais_mensais_dict.setdefault(mes_chave, OrderedDict())[unidade] = total["total_quantidade"]

    pedidos_agrupados = OrderedDict()
    agora = timezone.now()

    for pedido in pedidos_selecionados:
        artigo_nome = (pedido.artigo or "").lower()
        dias_antecedencia = 20 if "nobuck" in artigo_nome else 11

        pedido.previsao_inicio = pedido.dt_programada - timedelta(days=dias_antecedencia)
        prazo = pedido.dt_programada
        delta = prazo - agora

        if delta.total_seconds() < 0:
            atraso = abs(delta)
            if atraso.days >= 1:
                pedido.status_prazo_texto = f"Atrasado há {atraso.days} dia(s)"
            else:
                horas = int(atraso.total_seconds() // 3600)
                pedido.status_prazo_texto = f"Atrasado há {horas} hora(s)"
            pedido.status_prazo_classe = "atrasado"
        elif prazo.date() == agora.date():
            pedido.status_prazo_texto = "Vence hoje"
            pedido.status_prazo_classe = "hoje"
        else:
            dias_restantes = delta.days
            pedido.status_prazo_texto = f"Faltam {dias_restantes} dia(s)"
            pedido.status_prazo_classe = "adiantado"

        mes_chave = f"{pedido.dt_programada.year}-{pedido.dt_programada.month:02d}"
        if mes_chave not in pedidos_agrupados:
            pedidos_agrupados[mes_chave] = {
                "items": [],
                "subtotais": totais_mensais_dict.get(mes_chave, {}),
                "mes_display": f"{MESES_DO_ANO[pedido.dt_programada.month]} - {pedido.dt_programada.year}",
                "clientes": OrderedDict(),
            }

        pedidos_agrupados[mes_chave]["items"].append(pedido)
        nome_cliente = str(pedido.cliente)
        pedidos_agrupados[mes_chave]["clientes"].setdefault(nome_cliente, {})
        unidade = pedido.unidade_medida or "N/D"
        pedidos_agrupados[mes_chave]["clientes"][nome_cliente].setdefault(unidade, 0)
        pedidos_agrupados[mes_chave]["clientes"][nome_cliente][unidade] += float(pedido.quantidade or 0)

    for grupo in pedidos_agrupados.values():
        grupo["clientes"] = OrderedDict(sorted(grupo["clientes"].items(), key=lambda x: x[0].lower()))

    if output_format == "excel":
        return HttpResponse("Exportação excel não habilitada neste patch.", content_type="text/plain")

    context = {"pedidos_agrupados": pedidos_agrupados, "data_atual": agora}
    return render(request, "pedido/impressao_data.html", context)


def gerar_excel_data_view(pedidos_agrupados, totais_gerais):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pedidos por Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    group_font = Font(bold=True, size=12)
    total_font = Font(bold=True)

    headers = ["Contrato", "Cód. Interno", "Cliente", "Artigo", "Quantidade", "Unidade", "Prazo"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for _, dados_mes in pedidos_agrupados.items():
        sheet.append([])
        group_cell = sheet.cell(row=sheet.max_row, column=1, value=dados_mes["mes_display"])
        group_cell.font = group_font
        sheet.merge_cells(
            start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=len(headers)
        )

        for pedido in dados_mes["items"]:
            row_data = [
                pedido.nr_contract,
                pedido.nr_pedido_interno,
                pedido.cliente,
                pedido.artigo,
                pedido.quantidade,
                pedido.unidade_medida,
                pedido.dt_programada.strftime("%d/%m/%Y") if pedido.dt_programada else "N/A",
            ]
            sheet.append(row_data)

        for i, (unidade, total) in enumerate(dados_mes["subtotais"].items()):
            label = "Subtotais:" if i == 0 else ""
            sheet.append(["", "", "", label, total, f"({unidade})", ""])
            for cell in sheet[sheet.max_row]:
                cell.font = total_font

    sheet.append([])
    for i, (unidade, total) in enumerate(totais_gerais.items()):
        label = "TOTAIS GERAIS:" if i == 0 else ""
        sheet.append(["", "", "", label, total, f"({unidade})", ""])
        for cell in sheet[sheet.max_row]:
            cell.font = total_font
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for i, column_cells in enumerate(sheet.columns):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(i + 1)].width = max_length + 2

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# ==============================================================================
# 3) SIMULADOR (novo fluxo: romaneio -> pallets)
# ==============================================================================

@staff_member_required
def simuladorlogisticaview(request):
    veiculos = Veiculo.objects.all().select_related("tipo", "transportadora").order_by("placa")
    tipos_veiculo = TipoVeiculo.objects.all().order_by("nome")
    context = {
        "veiculos": veiculos,
        "tipos_veiculo": tipos_veiculo,
        "data_atual": timezone.now(),
        "pedidos": [],
    }
    return render(request, "pedido/simulador_logistica.html", context)


def _pes2_to_m2(pes2: float | None) -> float:
    if pes2 is None:
        return 0.0
    return round(float(pes2) / PES2_TO_M2_DIVISOR, 3)


SQL_ROMANEIO = r"""
SELECT TOP 100000
    Estoque_Expedicao.Cd_Romaneo_Faturamento AS cd_romaneio_faturamento,
    Estoque_Expedicao.Cd_Pedido_Comercial_Artigo AS seq_ped_entrega,
    Nr_Pallet AS nr_pallet,
    Nr_Pallet_Nr_Embalagem AS pallet,

    convert(char(80),
        case when SeA_Artigo.Nome is not null then
            SeA_Artigo.Nome +
            case when SeA_Artigo_Cor.Nome is not null then ' ' + SeA_Artigo_Cor.Nome else '' end +
            case when WB_SeA_Espessura.Nome is not null then ' ' + WB_SeA_Espessura.Nome else '' end +
            case when Estoque_Expedicao.Outros_Detalhes is not null then ' ' + rtrim(Estoque_Expedicao.Outros_Detalhes) else '' end +
            case when WB_Artigo.Unid_em_Couro = 1 then ' INTEIRO' else ' MEIO' end
        else
            WB_Artigo.Nome +
            case when WB_SeA_Espessura.Nome is not null then ' ' + WB_SeA_Espessura.Nome else '' end +
            case when WB_Artigo_Tipo.Nome is not null then ' ' + WB_Artigo_Tipo.Nome else '' end +
            case when WB_Artigo_Tamanho.Nome is not null then ' ' + WB_Artigo_Tamanho.Nome else '' end +
            case when Estoque_Expedicao.Outros_Detalhes is not null then ' ' + rtrim(Estoque_Expedicao.Outros_Detalhes) else '' end
        end
    ) as produto,

    Case when left(WB_Artigo.Produto,5) = 'RASPA' then
        Peso_Pallet_Lq
    else
        convert(numeric(18,2),Pes2 * (1-(isnull(Estoque_Expedicao.Perc_Desconto_Metragem,0)/ 100)) )
    end as pes2,

    (Pecas) as pecas,
    Peso_Pallet_Lq as peso_pallet_lq,

    Estoque_Expedicao_Romaneo.Dt_Saida as dt_saida,
    Cliente.Nome_Fantasia_CNPJ as cliente

FROM Estoque_Expedicao
INNER JOIN WB_Artigo ON Estoque_Expedicao.Cd_WB_Artigo = WB_Artigo.Codigo
INNER JOIN Estoque_Expedicao_Romaneo ON Estoque_Expedicao.Cd_Romaneo_Faturamento = Estoque_Expedicao_Romaneo.Codigo
INNER JOIN Fornecedor_Cliente_CNPJ Cliente ON Estoque_Expedicao_Romaneo.Cd_Endereco_Entrega = Cliente.Codigo
LEFT OUTER JOIN WB_SeA_Espessura ON WB_SeA_Espessura.Codigo = Estoque_Expedicao.Cd_WB_SeA_Espessura
LEFT OUTER JOIN WB_Artigo_Tipo ON WB_Artigo_Tipo.Codigo = Estoque_Expedicao.Cd_WB_Artigo_Tipo
LEFT OUTER JOIN WB_Artigo_Tamanho ON WB_Artigo_Tamanho.Codigo = Estoque_Expedicao.Cd_WB_Artigo_Tamanho
LEFT OUTER JOIN SeA_Artigo_Cor ON SeA_Artigo_Cor.Codigo = Estoque_Expedicao.Cd_SeA_Artigo_Cor
LEFT OUTER JOIN SeA_Artigo ON SeA_Artigo.Codigo = Estoque_Expedicao.Cd_SeA_Artigo
LEFT OUTER JOIN Unidade_de_Producao ON Estoque_Expedicao.Cd_Unidade_de_Producao = Unidade_de_Producao.Codigo
WHERE Estoque_Expedicao.Cd_Romaneo_Faturamento = %s
"""


def executar_query_romaneio(cd_romaneio: str) -> list[dict]:
    host = os.environ.get("MSSQL_HOST", "")
    port = int(os.environ.get("MSSQL_PORT", "1433"))
    user = os.environ.get("MSSQL_USER", "")
    password = os.environ.get("MSSQL_PASSWORD", "")
    database = os.environ.get("MSSQL_DB", "")

    if not all([host, user, password, database]):
        raise Exception("Credenciais MSSQL não configuradas (MSSQL_HOST/USER/PASSWORD/DB).")

    con = pymssql.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        as_dict=True,
        timeout=20,
        login_timeout=10,
        charset="UTF-8",
    )
    try:
        with con.cursor() as cursor:
            cursor.execute(SQL_ROMANEIO, (cd_romaneio,))
            rows = cursor.fetchall() or []

        out = []
        for r in rows:
            out.append({
                "cd_romaneio_faturamento": str(r.get("cd_romaneio_faturamento") or cd_romaneio),
                "seq_ped_entrega": str(r.get("seq_ped_entrega") or "").strip(),
                "nr_pallet": str(r.get("nr_pallet") or "").strip(),
                "pallet": str(r.get("pallet") or "").strip(),
                "produto": (r.get("produto") or "").strip() if isinstance(r.get("produto"), str) else r.get("produto"),
                "cliente": (r.get("cliente") or "").strip() if isinstance(r.get("cliente"), str) else r.get("cliente"),
                "pes2": float(r.get("pes2") or 0),
                "peso_pallet_lq": float(r.get("peso_pallet_lq") or 0),
                "pecas": int(r.get("pecas") or 0),
                "dt_saida": r.get("dt_saida"),
            })
        return out
    finally:
        con.close()


@staff_member_required
def api_romaneio_pallets(request):
    cd_romaneio = (request.GET.get("cd_romaneio") or "").strip()
    if not cd_romaneio:
        return JsonResponse({"status": "error", "message": "Informe cd_romaneio."}, status=400)

    try:
        linhas = executar_query_romaneio(cd_romaneio)
    except NotImplementedError as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=501)
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Erro ao consultar romaneio: {e}"}, status=500)

    itens = []
    for row in linhas:
        seq = str(row.get("seq_ped_entrega") or "").strip()
        nr_pallet = str(row.get("nr_pallet") or "").strip()
        pallet = str(row.get("pallet") or "").strip()

        pes2 = row.get("pes2")
        m2 = _pes2_to_m2(pes2)

        pedido = Pedido.objects.filter(nr_pedido_interno=seq).first()
        if not pedido:
            itens.append({
                "ok": False,
                "erro": f"Pedido não encontrado para Seq_Ped_Entrega={seq}",
                "seq_ped_entrega": seq,
                "nr_pallet": nr_pallet,
                "pallet": pallet,
                "pes2": pes2,
                "m2": m2,
            })
            continue

        itens.append({
            "ok": True,
            "pedido_id": pedido.id,
            "seq_ped_entrega": seq,
            "nr_contract": pedido.nr_contract,
            "nr_pedido_interno": pedido.nr_pedido_interno,
            "cliente": row.get("cliente") or pedido.cliente,
            "produto": row.get("produto") or pedido.produto,
            "cd_romaneio_faturamento": cd_romaneio,
            "nr_pallet": nr_pallet,
            "pallet": pallet,
            "pes2": pes2,
            "m2": m2,
            "peso_liquido": row.get("peso_pallet_lq"),
            "pecas": row.get("pecas"),
            "saldo_pedido_m2": pedido.saldo_a_entregar,
        })

    return JsonResponse({"status": "success", "cd_romaneio": cd_romaneio, "items": itens})


# ==============================================================================
# 4) SALVAR LAYOUT (cria Embarque + Itens, valida saldo, trava duplicidade)
# ==============================================================================

@csrf_exempt
@staff_member_required
def salvarlayoutlogistica(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Método inválido."}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"status": "error", "message": "JSON inválido."}, status=400)

    cd_romaneio = str(payload.get("cd_romaneio_faturamento") or "").strip()
    veiculo_id = payload.get("veiculo_id")
    tipo_veiculo_id = payload.get("tipo_veiculo_id")
    data_embarque = payload.get("data_embarque")
    itens = payload.get("itens") or []
    mapa = payload.get("mapa_carga_json")

    if not cd_romaneio:
        return JsonResponse({"status": "error", "message": "cd_romaneio_faturamento é obrigatório."}, status=400)
    if not data_embarque:
        return JsonResponse({"status": "error", "message": "data_embarque é obrigatório (YYYY-MM-DD)."}, status=400)
    if not isinstance(itens, list) or not itens:
        return JsonResponse({"status": "error", "message": "itens é obrigatório (lista)."}, status=400)

    def resolver_veiculo_obrigatorio():
        if veiculo_id:
            return Veiculo.objects.select_related("tipo", "transportadora").get(pk=veiculo_id)

        if not tipo_veiculo_id:
            raise ValidationError("Selecione um veículo cadastrado OU um tipo de veículo (padrão).")

        tipo = TipoVeiculo.objects.get(pk=tipo_veiculo_id)

        transp, _ = Transportadora.objects.get_or_create(
            nome="NÃO INFORMADA",
            defaults={"cnpj": None, "contato": None},
        )

        placa_generica = f"GEN-{tipo.nome.upper()[:7]}"
        v, _ = Veiculo.objects.get_or_create(
            placa=placa_generica,
            defaults={"transportadora": transp, "tipo": tipo, "motorista": None},
        )

        changed = False
        if v.tipo_id != tipo.id:
            v.tipo = tipo
            changed = True
        if v.transportadora_id != transp.id:
            v.transportadora = transp
            changed = True
        if changed:
            v.save(update_fields=["tipo", "transportadora"])

        return v

    with transaction.atomic():
        try:
            veiculo = resolver_veiculo_obrigatorio()
        except (Veiculo.DoesNotExist, TipoVeiculo.DoesNotExist) as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        except ValidationError as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

        embarque = Embarque.objects.create(
            veiculo=veiculo,
            data_embarque=data_embarque,
            mapa_carga_json=mapa,
            finalizado=False,
        )

        criados = 0
        for it in itens:
            seq = str(it.get("seq_ped_entrega") or "").strip()
            nr_pallet = str(it.get("nr_pallet") or "").strip()
            pallet = str(it.get("pallet") or "").strip()
            pes2 = it.get("pes2")

            if not seq or not nr_pallet or not pallet:
                raise ValidationError("Item inválido: seq_ped_entrega, nr_pallet e pallet são obrigatórios.")

            pedido = Pedido.objects.get(nr_pedido_interno=seq)

            m2 = _pes2_to_m2(pes2)
            if m2 <= 0:
                raise ValidationError(f"m² inválido para pallet {nr_pallet}/{pallet} (Pes2={pes2}).")

            # Tolerância de 50% sobre o saldo para cobrir variações do romaneio real
            tolerancia = 0.50
            saldo_permitido = float(pedido.saldo_a_entregar or 0) * (1 + tolerancia)
            if m2 > saldo_permitido:
                raise ValidationError(
                    f"Pallet {nr_pallet}/{pallet}: m² ({m2:.2f}) excede "
                    f"saldo ({pedido.saldo_a_entregar:.2f}) +50% = {saldo_permitido:.2f}."
                )

            obj = ItemEmbarque(
                embarque=embarque,
                pedido=pedido,
                metragem_embarcada=m2,
                cd_romaneio_faturamento=cd_romaneio,
                nr_pallet=nr_pallet,
                pallet=pallet,
                pes2_original=pes2,
                peso_liquido=it.get("peso_liquido"),
                pecas=it.get("pecas"),
                pos_x=float(it.get("pos_x") or 0),
                pos_y=float(it.get("pos_y") or 0),
                pos_z=int(it.get("pos_z") or 0),
                largura=float(it.get("largura") or 1.2),
                comprimento=float(it.get("comprimento") or 1.0),
                altura=float(it.get("altura") or 1.0),
                rotacionado=bool(it.get("rotacionado") or False),
            )

            obj.full_clean()
            try:
                obj.save()
            except IntegrityError:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": f"Pallet já embarcado: romaneio={cd_romaneio}, nr_pallet={nr_pallet}, pallet={pallet}",
                    },
                    status=409,
                )

            criados += 1

    return JsonResponse(
        {"status": "success", "message": "Carga salva!", "embarque_id": embarque.id, "itens_criados": criados}
    )



@staff_member_required
def embarques_list_view(request):
    qs = (
        Embarque.objects
        .select_related("veiculo", "veiculo__tipo", "veiculo__transportadora")
        .annotate(total_m2=Sum("itens_embarcados__metragem_embarcada"))
        .order_by("-data_embarque", "-id")
    )

    dt_ini = (request.GET.get("dt_ini") or "").strip()
    dt_fim = (request.GET.get("dt_fim") or "").strip()
    finalizado = (request.GET.get("finalizado") or "").strip()

    if dt_ini:
        qs = qs.filter(data_embarque__gte=dt_ini)
    if dt_fim:
        qs = qs.filter(data_embarque__lte=dt_fim)
    if finalizado in ("0", "1"):
        qs = qs.filter(finalizado=(finalizado == "1"))

    return render(request, "pedido/embarques_list.html", {"embarques": qs})


@staff_member_required
def embarque_detail_view(request, pk: int):
    embarque = (
        Embarque.objects
        .select_related("veiculo", "veiculo__tipo", "veiculo__transportadora")
        .prefetch_related("itens_embarcados__pedido")
        .annotate(
            total_m2=Sum("itens_embarcados__metragem_embarcada"),
            total_pes2=Sum("itens_embarcados__pes2_original")
        )
        .get(pk=pk)
    )

    itens = embarque.itens_embarcados.all().order_by("pos_z", "pos_y", "pos_x", "id")
    return render(request, "pedido/embarque_detail.html", {"embarque": embarque, "itens": itens})


@staff_member_required
def embarque_print_view(request, pk: int):
    embarque = get_object_or_404(
        Embarque.objects
        .select_related("veiculo", "veiculo__tipo", "veiculo__transportadora")
        .prefetch_related("itens_embarcados__pedido")
        .annotate(
            total_m2=Sum("itens_embarcados__metragem_embarcada"),
            total_pes2=Sum("itens_embarcados__pes2_original")
        ),
        pk=pk,
    )
    
    itens = embarque.itens_embarcados.all().order_by("pos_z", "pos_y", "pos_x", "id")
    return render(request, "pedido/embarque_print.html", {"embarque": embarque, "itens": itens})


from .models import ReservaEmbarque  # adicione no import de models

@csrf_exempt
@require_POST
def iniciar_logistica_previsao_view(request):
    """
    Inicia uma logística (Embarque) com base em data + caminhão + lista de pedidos.

    JSON:
    {
      "data_embarque": "2026-04-10",
      "veiculo_id": 1,              # OU
      "tipo_veiculo_id": 2,         # se não tiver veiculo_id
      "itens": [
        { "cd_pedido": 123, "m2_previsto": 100.0 },
        ...
      ]
    }
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"status": "error", "message": "JSON inválido."}, status=400)

    data_embarque_str = payload.get("data_embarque")
    veiculo_id = payload.get("veiculo_id")
    tipo_veiculo_id = payload.get("tipo_veiculo_id")
    itens = payload.get("itens") or []

    if not data_embarque_str:
        return JsonResponse({"status": "error", "message": "data_embarque é obrigatória (YYYY-MM-DD)."}, status=400)

    data_embarque = parse_date(data_embarque_str)
    if not data_embarque:
        return JsonResponse({"status": "error", "message": "data_embarque inválida."}, status=400)

    if not isinstance(itens, list) or not itens:
        return JsonResponse({"status": "error", "message": "itens é obrigatório (lista)."}, status=400)

    def resolver_veiculo_obrigatorio():
        if veiculo_id:
            return Veiculo.objects.select_related("tipo", "transportadora").get(pk=veiculo_id)

        if not tipo_veiculo_id:
            raise ValidationError("Selecione um veículo cadastrado OU um tipo (padrão).")

        tipo = TipoVeiculo.objects.get(pk=tipo_veiculo_id)

        transp, _ = Transportadora.objects.get_or_create(
            nome="NÃO INFORMADA",
            defaults={"cnpj": None, "contato": None},
        )

        placa_generica = f"GEN-{tipo.nome.upper()[:7]}"
        v, _ = Veiculo.objects.get_or_create(
            placa=placa_generica,
            defaults={"transportadora": transp, "tipo": tipo, "motorista": None},
        )

        changed = False
        if v.tipo_id != tipo.id:
            v.tipo = tipo
            changed = True
        if v.transportadora_id != transp.id:
            v.transportadora = transp
            changed = True
        if changed:
            v.save(update_fields=["tipo", "transportadora"])

        return v

    with transaction.atomic():
        try:
            veiculo = resolver_veiculo_obrigatorio()
        except (Veiculo.DoesNotExist, TipoVeiculo.DoesNotExist) as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        except ValidationError as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

        embarque = Embarque.objects.create(
            veiculo=veiculo,
            data_embarque=data_embarque,
            mapa_carga_json=None,
            finalizado=False,
        )

        criados = 0
        for it in itens:
            nr_contract = (it.get("nr_contract") or "").strip()
            nr_pedido_interno = (it.get("nr_pedido_interno") or "").strip()
            m2_prev = float(it.get("m2_previsto") or 0)

            if not nr_contract and not nr_pedido_interno:
                raise ValidationError(
                    "Item inválido: informe ao menos nr_contract ou nr_pedido_interno."
                )
            if m2_prev <= 0:
                raise ValidationError("Metragem prevista deve ser maior que zero.")

            qs = Pedido.objects.all()
            if nr_contract:
                qs = qs.filter(nr_contract=nr_contract)
            if nr_pedido_interno:
                qs = qs.filter(nr_pedido_interno=nr_pedido_interno)

            pedido = qs.first()
            if not pedido:
                raise ValidationError(
                    f"Pedido não encontrado para nr_contract={nr_contract or '∅'}, "
                    f"nr_pedido_interno={nr_pedido_interno or '∅'}."
                )

            reserva = ReservaEmbarque(
                embarque=embarque,
                pedido=pedido,
                metragem_prevista_m2=m2_prev,
            )
            reserva.full_clean()
            reserva.save()
            criados += 1

    return JsonResponse(
        {
            "status": "success",
            "message": "Logística iniciada (previsão criada).",
            "embarque_id": embarque.id,
            "reservas_criadas": criados,
        }
    )



@csrf_exempt
@staff_member_required
@require_GET
def listar_embarques_previstos_view(request):
    """
    Lista embarques (abertos ou todos) por data e filtro de texto,
    para o simulador continuar uma previsão existente.

    GET params:
      data = YYYY-MM-DD (obrigatório)
      q    = texto livre (placa, cidade, estado, transportadora) opcional
    Retorna JSON com embarques resumidos.
    """
    data_str = request.GET.get("data") or ""
    q = (request.GET.get("q") or "").strip()

    if not data_str:
        return JsonResponse({"status": "error", "message": "Parâmetro 'data' é obrigatório (YYYY-MM-DD)."}, status=400)

    try:
        data = parse_date(data_str)
    except Exception:
        data = None
    if not data:
        return JsonResponse({"status": "error", "message": "Data inválida."}, status=400)

    qs = (
        Embarque.objects
        .select_related("veiculo", "veiculo__tipo", "veiculo__transportadora")
        .filter(data_embarque=data)
        .order_by("veiculo__transportadora__nome", "veiculo__placa")
    )

    # opcional: mostrar só não finalizados
    # qs = qs.filter(finalizado=False)

    if q:
        qs = qs.filter(
            Q(veiculo__placa__icontains=q) |
            Q(veiculo__transportadora__nome__icontains=q)
            # se tiver campos de cidade/estado no modelo de Transportadora, adiciona aqui
            # Q(veiculo__transportadora__cidade__icontains=q) |
            # Q(veiculo__transportadora__estado__icontains=q)
        )

    embarques_data = []
    for e in qs[:50]:
        embarques_data.append({
            "id": e.id,
            "data": e.data_embarque.isoformat(),
            "placa": e.veiculo.placa,
            "tipo": e.veiculo.tipo.nome,
            "transportadora": e.veiculo.transportadora.nome,
            "finalizado": bool(e.finalizado),
        })

    return JsonResponse({"status": "success", "items": embarques_data})

@csrf_exempt
@staff_member_required
@require_POST
def salvar_previsao_logistica_view(request):
    """
    Salva uma PrevisaoEmbarque + PrevisaoItemEmbarque.

    JSON esperado:
    {
        "data_prevista": "2026-04-01",
        "observacao": "Texto livre",
        "itens": [
            {
                "pedido_id": 1,
                "m2": 120.5,
                "cd_romaneio_faturamento": "123456"  // opcional
            },
            ...
        ]
    }
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"status": "error", "message": "JSON inválido."}, status=400)

    data_prevista_str = payload.get("data_prevista")
    observacao = (payload.get("observacao") or "").strip()
    itens = payload.get("itens") or []

    if not data_prevista_str:
        return JsonResponse({"status": "error", "message": "data_prevista é obrigatória (YYYY-MM-DD)."}, status=400)

    data_prevista = parse_date(data_prevista_str)
    if not data_prevista:
        return JsonResponse({"status": "error", "message": "data_prevista inválida."}, status=400)

    if not isinstance(itens, list) or not itens:
        return JsonResponse({"status": "error", "message": "itens é obrigatório (lista)."}, status=400)

    with transaction.atomic():
        previsao = PrevisaoEmbarque.objects.create(
            data_prevista=data_prevista,
            observacao=observacao or None,
        )

        criados = 0
        for it in itens:
            pedido_id = it.get("pedido_id")
            m2 = float(it.get("m2") or 0)
            cd_romaneio = (it.get("cd_romaneio_faturamento") or "").strip()

            if not pedido_id:
                raise ValidationError("Item inválido: pedido_id é obrigatório.")
            if m2 <= 0:
                raise ValidationError("Metragem prevista deve ser maior que zero.")

            pedido = Pedido.objects.get(pk=pedido_id)

            prev_item = PrevisaoItemEmbarque(
                previsao=previsao,
                pedido=pedido,
                metragem_prevista_m2=m2,
                cd_romaneio_faturamento=cd_romaneio or None,
            )
            prev_item.full_clean()
            prev_item.save()
            criados += 1

    return JsonResponse(
        {
            "status": "success",
            "message": "Previsão salva.",
            "previsao_id": previsao.id,
            "itens_criados": criados,
        }
    )


@csrf_exempt
@staff_member_required
@require_POST
def converter_previsao_para_embarque_view(request):
    """
    Converte uma PrevisaoEmbarque em Embarque real.

    JSON esperado:
    {
        "previsao_id": 123,
        "veiculo_id": 1,          # opcional
        "tipo_veiculo_id": 2,     # opcional (fallback se não vier veiculo_id)
        "data_embarque": "2026-04-01"
    }
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"status": "error", "message": "JSON inválido."}, status=400)

    previsao_id = payload.get("previsao_id")
    veiculo_id = payload.get("veiculo_id")
    tipo_veiculo_id = payload.get("tipo_veiculo_id")
    data_embarque_str = payload.get("data_embarque")

    if not previsao_id:
        return JsonResponse({"status": "error", "message": "previsao_id é obrigatório."}, status=400)

    if not data_embarque_str:
        return JsonResponse({"status": "error", "message": "data_embarque é obrigatório (YYYY-MM-DD)."}, status=400)

    data_embarque = parse_date(data_embarque_str)
    if not data_embarque:
        return JsonResponse({"status": "error", "message": "data_embarque inválida."}, status=400)

    try:
        previsao = (
            PrevisaoEmbarque.objects
            .prefetch_related("itens__pedido")
            .get(pk=previsao_id)
        )
    except PrevisaoEmbarque.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Previsão não encontrada."}, status=404)

    if previsao.convertido:
        return JsonResponse({"status": "error", "message": "Esta previsão já foi convertida em embarque."}, status=400)

    def resolver_veiculo_obrigatorio():
        if veiculo_id:
            return Veiculo.objects.select_related("tipo", "transportadora").get(pk=veiculo_id)

        if not tipo_veiculo_id:
            raise ValidationError("Selecione um veículo cadastrado OU um tipo de veículo (padrão).")

        tipo = TipoVeiculo.objects.get(pk=tipo_veiculo_id)

        transp, _ = Transportadora.objects.get_or_create(
            nome="NÃO INFORMADA",
            defaults={"cnpj": None, "contato": None},
        )

        placa_generica = f"GEN-{tipo.nome.upper()[:7]}"
        v, _ = Veiculo.objects.get_or_create(
            placa=placa_generica,
            defaults={"transportadora": transp, "tipo": tipo, "motorista": None},
        )

        changed = False
        if v.tipo_id != tipo.id:
            v.tipo = tipo
            changed = True
        if v.transportadora_id != transp.id:
            v.transportadora = transp
            changed = True
        if changed:
            v.save(update_fields=["tipo", "transportadora"])

        return v

    with transaction.atomic():
        try:
            veiculo = resolver_veiculo_obrigatorio()
        except (Veiculo.DoesNotExist, TipoVeiculo.DoesNotExist) as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        except ValidationError as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

        embarque = Embarque.objects.create(
            veiculo=veiculo,
            data_embarque=data_embarque,
            mapa_carga_json=None,
            finalizado=False,
        )

        criados = 0

        for prev_item in previsao.itens.all():
            pedido = prev_item.pedido
            m2 = float(prev_item.metragem_prevista_m2 or 0)

            if m2 <= 0:
                raise ValidationError(
                    f"Metragem prevista inválida para pedido {pedido.nr_contract}."
                )

            # Usa mesma filosofia de tolerância do salvarlayoutlogistica
            tolerancia = 0.50
            saldo_permitido = float(pedido.saldo_a_entregar or 0) * (1 + tolerancia)
            if m2 > saldo_permitido + 1e-9:
                raise ValidationError(
                    f"Previsão do pedido {pedido.nr_contract}: m² ({m2:.2f}) excede "
                    f"saldo ({pedido.saldo_a_entregar:.2f}) +50% = {saldo_permitido:.2f}."
                )

            # Ainda não temos romaneio/pallet real → gera identificadores sintéticos
            obj = ItemEmbarque(
                embarque=embarque,
                pedido=pedido,
                metragem_embarcada=m2,
                cd_romaneio_faturamento=(
                    prev_item.cd_romaneio_faturamento or f"PREV-{previsao.id}"
                ),
                nr_pallet=f"PREV-{previsao.id}-{prev_item.id}",
                pallet=f"PREV-{previsao.id}-{prev_item.id}",
                pes2_original=None,
                peso_liquido=None,
                pecas=None,
                pos_x=0,
                pos_y=0,
                pos_z=0,
                largura=1.2,
                comprimento=1.0,
                altura=1.0,
                rotacionado=False,
            )

            obj.full_clean()
            obj.save()
            criados += 1

        previsao.convertido = True
        previsao.save(update_fields=["convertido"])

    return JsonResponse(
        {
            "status": "success",
            "message": "Previsão convertida em embarque.",
            "embarque_id": embarque.id,
            "itens_criados": criados,
        }
    )


# ==========================================================
# COMPATIBILIDADE: manter nomenclaturas antigas (core urls)
# ==========================================================

imprimir_pedido_view = imprimirpedidoview
imprimir_pedido_data_view = imprimirpedidodataview
simulador_logistica_view = simuladorlogisticaview
salvar_layout_logistica = salvarlayoutlogistica
converter_previsao_para_embarque = converter_previsao_para_embarque_view